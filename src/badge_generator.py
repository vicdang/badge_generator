# -*- coding: utf-8 -*-
"""
Badge Generator - Create ID badges with automatic face detection

Authors: Vic Dang
Usage example:
   python execute.py -d -v exec
"""

import argparse
import logging
import logging.config
import os
import re
import sys
import time
import unicodedata
from argparse import ArgumentDefaultsHelpFormatter as Formatter
from pathlib import Path
from typing import Tuple, Optional, List, Dict
from pathlib import Path
import sys

# Add parent directory to path to import tools package
sys.path.insert(0, str(Path(__file__).parent.parent))

import cv2
import numpy as np
import qrcode
from PIL import Image, ImageDraw, ImageFont

from tools.util import Utilities as uT
from tools.image_manager import ImageManager


# Constants
CUR_PATH = Path(__file__).parent.parent.resolve()
CONF_DIR = CUR_PATH / 'config'
CONFIG_FILE = CONF_DIR / 'config.ini'
LOGGER = logging.getLogger(__name__)

# Configuration constants
FACE_CASCADE_PATH = "resources/haar_cascade/haarcascade_frontalface_default.xml"
ORIENTATION_TAG = 274
TEXT_RENDER_CONFIGS = {
    'username': {'reduction': 0},
    'position': {'reduction': 10},
    'userid': {'reduction': 15},
}
MIN_TEXT_WIDTH = 50
FONT_SIZE_DECREMENT = 5
DEBUG_IMAGE_RESIZE_HEIGHT = 1024




# Initialize config after Utility class is defined
config = None
try:
    config = Utility.get_config()
except Exception as e:
    # Fallback: config will be loaded in main if needed, but this prevents early failure
    pass


class ImageSizeException(Exception):
    """Raised when image is too small for processing"""
    pass


class UserInfoException(Exception):
    """Raised when user information parsing fails"""
    pass


class Utility:
    """Utility functions for badge generation"""

    @staticmethod
    def get_config():
        """Load configuration from config.ini"""
        try:
            from configparser import ConfigParser
        except ImportError:
            from ConfigParser import ConfigParser  # ver. < 3.0
        
        conf = ConfigParser()
        conf.read(CONFIG_FILE)
        return conf

    @staticmethod
    def validate(string: str, regex: str) -> str:
        """
        Validate a string against a regex pattern
        
        :param string: input string
        :param regex: regular expression to verify the string
        :return: string if valid, empty string otherwise
        """
        if re.search(regex, string):
            return string
        
        LOGGER.error(f"[{string}] does not match [{regex}]")
        return ""

    @staticmethod
    def convert_img(input_img: str) -> np.ndarray:
        """
        Convert image file to OpenCV format
        
        :param input_img: path to input image
        :return: image as numpy array
        """
        return cv2.imdecode(np.fromfile(input_img, dtype=np.uint8), -1)

    @staticmethod
    def countdown(due_time: int) -> None:
        """
        Display countdown timer
        
        :param due_time: countdown duration in seconds
        """
        LOGGER.debug(f"Counting down for {due_time} seconds")
        while due_time:
            minute, second = divmod(due_time, 60)
            time_format = f'{minute:02d}:{second:02d}'
            print(time_format, end='\r')
            time.sleep(1)
            due_time -= 1

    @staticmethod
    def check_folder(folders: List[str]) -> None:
        """
        Prepare required folders
        
        :param folders: list of folder paths to create
        """
        for folder in folders:
            if not os.path.isdir(folder):
                LOGGER.debug(f"Creating directory: {folder}")
                os.makedirs(folder, exist_ok=True)

    @staticmethod
    def resize_with_aspect_ratio(
        image: np.ndarray,
        width: Optional[int] = None,
        height: Optional[int] = None,
        inter: int = cv2.INTER_AREA
    ) -> np.ndarray:
        """
        Resize image while maintaining aspect ratio
        
        :param image: input image
        :param width: target width (optional)
        :param height: target height (optional)
        :param inter: interpolation method
        :return: resized image
        """
        if width is None and height is None:
            return image
        
        h, w = image.shape[:2]
        
        if width is None:
            r = height / float(h)
            dim = (int(w * r), height)
        else:
            r = width / float(w)
            dim = (width, int(h * r))

        return cv2.resize(image, dim, interpolation=inter)


class ImageMaker:
    """Badge image generator with face detection and text rendering"""

    def __init__(self, name: str, arg: argparse.Namespace, conf) -> None:
        self.arg = arg
        self.name = name
        self.src_path = CUR_PATH / arg.src_path
        self.des_path = CUR_PATH / arg.des_path
        self.tmp_path = CUR_PATH / conf.get("general", "tmp_path")
        self.template = CUR_PATH / arg.template
        self.debug = arg.debug or False
        self.img_prefix = conf.get("general", "img_prefix")
        self.tpl_avatar_x = conf.getint("template", "avata_x")
        self.tpl_avatar_y = conf.getint("template", "avata_y")
        self.tpl_avatar_w = conf.getint("template", "avata_w")
        self.tpl_avatar_h = conf.getint("template", "avata_h")
        self.tpl_w = conf.getint("template", "width")
        self.tpl_h = conf.getint("template", "height")
        self.conf = conf
        self.base_text_size = self.conf.getint("general", "text_size")
        self.positions = uT.get_dict_positions()
        self.re_name = r"^[\w\.\- ]+$"
        self.re_id = r"^[\w]?[\d]+$"

        self.img = self.img_resized = self.img_cropped = self.bg_img = None
        self.user_pos = self.positions.get("E", "")
        self.user_name = self.user_id = ""
        self.img_num = 1

    def get_focus_position(self, input_img: str) -> Tuple[int, int]:
        """
        Detect faces in the image and return the focus position
        
        :param input_img: path to the input image
        :return: Focus position coordinates (x, y)
        """
        face_cascade = cv2.CascadeClassifier(FACE_CASCADE_PATH)
        image = Utility.convert_img(input_img)
        gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
        scale_factor = self.conf.getfloat("avata", "scalefactor")
        faces = face_cascade.detectMultiScale(
            gray_image,
            scaleFactor=scale_factor,
            minNeighbors=3,
            minSize=(30, 30),
            maxSize=(200, 200),
            flags=cv2.CASCADE_SCALE_IMAGE
        )

        LOGGER.info(f"Found {len(faces)} faces!")
        x = y = 0
        
        if len(faces):
            a, b, c, d = faces[0]
            for i, j, w, h in faces:
                LOGGER.debug(f"Detected face: [{i},{j},{w},{h}]")
                c = int((c + w) / 2)
                d = int((d + h) / 2)
                a = int((a + i) / 2)
                b = int((b + j) / 2)
                x = a + int(c / 2)
                y = b + int(d / 2)
                
                if self.arg.debug:
                    cv2.rectangle(
                        image,
                        (i + int(w / 2), j + int(h / 2)),
                        (i + int(w / 2) + 2, j + int(h / 2) + 2),
                        (0, 200, 100), 2
                    )
                    cv2.rectangle(image, (i, j), (i + w, j + h), (0, 200, 100), 2)
                    cv2.rectangle(image, (x, y), (x + 2, y + 2), (0, 255, 255), 2)
        
        if self.debug:
            resize = Utility.resize_with_aspect_ratio(image, height=DEBUG_IMAGE_RESIZE_HEIGHT)
            cv2.imshow("Faces found", resize)
            cv2.imwrite(str(self.tmp_path / f"faces_{self.name}"), image)
            if self.arg.test and self.arg.verbose:
                cv2.waitKey(0)
        
        LOGGER.debug(f"Focus coordinate: [{x},{y}]")
        return x, y

    @staticmethod
    def convert_images(src_path: str, des_path: str) -> None:
        """
        Convert images to PNG format
        
        :param src_path: source directory path
        :param des_path: destination directory path
        """
        format_list = uT.get_list_file_extensions()
        LOGGER.debug(f"Source files: {os.listdir(src_path)}")
        
        files = [
            f for f in os.listdir(src_path)
            if os.path.isfile(os.path.join(src_path, f))
            and any(ext in f for ext in format_list)
        ]
        LOGGER.debug(f"Files to convert: {files}")
        
        for file in files:
            name = file.split(".")[0]
            file_path = os.path.join(src_path, file)
            des_file_path = os.path.join(des_path, f"{name}.png")
            
            LOGGER.info(f"Converting {file_path} to PNG format...")
            im = Image.open(file_path).convert("RGBA")
            im.save(des_file_path, format="png")

    def correct_img(self) -> None:
        """Correct image orientation based on EXIF data"""
        try:
            exif = self.img._getexif()
            if exif:
                exif = dict(exif.items())
                if exif.get(ORIENTATION_TAG) == 3:
                    self.img = self.img.rotate(180, expand=True)
                elif exif.get(ORIENTATION_TAG) == 6:
                    self.img = self.img.rotate(270, expand=True)
                elif exif.get(ORIENTATION_TAG) == 8:
                    self.img = self.img.rotate(90, expand=True)
        except (AttributeError, KeyError) as err:
            LOGGER.error(f"Error correcting image orientation: {err}")

    def resize_img(self) -> None:
        """Resize image to match template avatar dimensions"""
        img_w, img_h = self.img.size
        LOGGER.info(f"Width: {img_w}, Height: {img_h}")
        
        basewidth = self.tpl_avatar_w + self.conf.getint("template", "padding")
        
        if img_h >= img_w:
            wpercent = basewidth / float(img_w)
            hsize = int(float(img_h) * wpercent)
            self.img_resized = self.img.resize((basewidth, hsize), Image.LANCZOS)
        else:
            hpercent = basewidth / float(img_h)
            wsize = int(float(img_w) * hpercent)
            self.img_resized = self.img.resize((wsize, basewidth), Image.LANCZOS)
        
        LOGGER.debug(f"Resized image: {self.img_resized.size[0]} x {self.img_resized.size[1]}")
        self.img_resized.save(str(self.tmp_path / self.name), format="png")

    def crop_img(self) -> None:
        """Crop image around detected face"""
        basewidth = self.tpl_avatar_w + self.conf.getint("template", "padding")
        x, y = self.get_focus_position(str(self.tmp_path / self.name))
        img_r_w, img_r_h = self.img_resized.size
        
        if img_r_w < basewidth or img_r_h < basewidth:
            raise ImageSizeException("Image is too small, please try another.")
        
        base_w = basewidth / 2
        
        if x <= 0 and y <= 0:
            x = int(img_r_w / 2)
            y = int(img_r_h / 2)
        else:
            x = max(base_w, min(x, img_r_w - base_w))
            y = max(base_w, min(y, img_r_h - base_w))
        
        correct_x, correct_y = int(x - base_w), int(y - base_w)
        correct_w, correct_h = int(basewidth + correct_x), int(basewidth + correct_y)
        
        LOGGER.debug(
            f"Crop params: x:{x}, y:{y} - c_x:{correct_x}, c_y:{correct_y} - w:{basewidth}, h:{basewidth}"
        )
        
        draw = ImageDraw.Draw(self.img_resized)
        if self.debug:
            draw.rectangle(
                [correct_x, correct_y, correct_w, correct_h],
                width=3,
                outline="#0000ff"
            )
        
        self.img_cropped = self.img_resized.crop(
            (correct_x, correct_y, correct_w, correct_h)
        )
        LOGGER.debug(f"Cropped image: {self.img_cropped.size[0]} x {self.img_cropped.size[1]}")
        
        if self.arg.debug or self.arg.test:
            self.img_cropped.save(str(self.tmp_path / f"cr_{self.name}"), format="png")

    def parse_user_info(self) -> None:
        """Parse user information from filename"""
        img_info_arr = self.name.split(".")[0].split("_")
        
        self.user_name = Utility.validate(img_info_arr[0].strip(), self.re_name)
        if not self.user_name:
            raise UserInfoException("User name not found!")
        
        self.user_pos = img_info_arr[2].capitalize()
        if self.user_pos.strip().upper() in self.positions:
            self.user_pos = self.positions[self.user_pos.strip().upper()]
            LOGGER.info(f"Position: {self.user_pos}")
        elif self.user_pos.strip().upper() in [str(v).upper() for v in self.positions.values()]:
            self.user_pos = self.user_pos.strip().upper()
            LOGGER.info(f"Position: {self.user_pos}")
        else:
            msg = f"Position [{self.user_pos}] not in {list(self.positions.keys())}"
            LOGGER.error(msg)
            raise UserInfoException("User position is incorrect!")
        
        self.user_id = Utility.validate(img_info_arr[1].strip(), self.re_id)
        if not self.user_id:
            raise UserInfoException("User ID not found!")
        
        if len(img_info_arr) == 4:
            self.img_num = img_info_arr[3]

    def parse_qr_code(self) -> None:
        """Generate and paste QR code onto background image"""
        qr = qrcode.QRCode(
            version=self.conf.getint("qrcode", "version"),
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=self.conf.getint("qrcode", "boxsize"),
            border=self.conf.getint("qrcode", "border")
        )
        
        name = unicodedata.normalize('NFKD', self.user_name).encode('ascii', 'ignore')
        img_info = (
            f"Fullname: {name}, Position: {self.user_pos}, "
            f"Badge_Id: {self.user_id}, Company: https://www.tma.vn"
        )
        LOGGER.info(f"QR Code info: {img_info}")
        
        if self.arg.qr_text:
            qr_img = qrcode.make(self.arg.qr_text)
        else:
            qr.add_data(img_info)
            qr.make(fit=self.conf.get("qrcode", "fit"))
            qr_img = qr.make_image(
                fill_color=self.conf.get("qrcode", "fillcolor"),
                back_color=self.conf.get("qrcode", "backcolor")
            )
        
        qrx = self.conf.getint("qrcode", "qrx")
        qry = self.conf.getint("qrcode", "qry")
        
        # For RGB images, convert QR code to RGB and paste directly
        if qr_img.mode == 'RGBA':
            qr_rgb = Image.new('RGB', qr_img.size, 'white')
            qr_rgb.paste(qr_img, mask=qr_img.split()[3])
            qr_img = qr_rgb
        elif qr_img.mode != 'RGB':
            qr_img = qr_img.convert('RGB')
        
        # Paste QR code without mask parameter
        self.bg_img.paste(qr_img, (qrx, qry))

    def _render_text(
        self,
        draw: ImageDraw.ImageDraw,
        text: str,
        section: str,
        curr_y: float
    ) -> Tuple[int, int, float]:
        """
        Render text on badge with automatic sizing
        
        :param draw: ImageDraw object
        :param text: text to render
        :param section: config section (username, position, userid)
        :param curr_y: current Y position
        :return: (text_width, text_height, new_y_position)
        """
        if not text:
            return 0, 0, curr_y
        
        curr_y += self.conf.getint(section, "toppad")
        
        if self.arg.no_auto_size:
            size = self.conf.getint(section, "size")
        else:
            size = self.base_text_size - TEXT_RENDER_CONFIGS.get(section, {}).get('reduction', 0)
        
        font_path = CUR_PATH / self.conf.get(section, "font")
        msg = text.upper().strip() if section == "username" else text.strip()
        
        # Find optimal font size
        n_size = size
        while n_size > 0:
            font = ImageFont.truetype(str(font_path), n_size)
            bbox = draw.textbbox((0, 0), msg, font=font)
            tw = bbox[2] - bbox[0]
            
            LOGGER.info(f"{section}: {msg} width:{tw} height:{bbox[3] - bbox[1]}")
            
            if tw < (self.tpl_w - MIN_TEXT_WIDTH):
                break
            n_size -= FONT_SIZE_DECREMENT
        
        font = ImageFont.truetype(str(font_path), n_size)
        bbox = draw.textbbox((0, 0), msg, font=font)
        tw = bbox[2] - bbox[0]
        th = bbox[3] - bbox[1]
        
        # Ensure coordinates are integers
        text_x = int((self.tpl_w - tw) / 2)
        text_y = int(curr_y)
        
        try:
            draw.text(
                (text_x, text_y),
                msg,
                fill=self.conf.get(section, "color"),
                font=font
            )
        except Exception as e:
            LOGGER.warning(f"Text rendering for {section} at ({text_x},{text_y}): {e}")
            # Try with default black color
            try:
                draw.text((text_x, text_y), msg, fill=(0, 0, 0), font=font)
            except Exception as e2:
                LOGGER.warning(f"Fallback also failed: {e2}")
        
        return tw, th, curr_y + th

    def execute(self) -> None:
        """Execute badge generation process"""
        try:
            LOGGER.info(f"Processing image: {self.name}")
            
            template_img = Image.open(str(self.template), 'r')
            tpl_w, tpl_h = template_img.size
            
            # Update instance variables to match actual template size
            self.tpl_w = tpl_w
            self.tpl_h = tpl_h

            self.img = Image.open(str(self.src_path / self.name), 'r').convert("RGBA")
            
            self.correct_img()
            self.resize_img()
            self.crop_img()
            
            # Create background
            self.bg_img = Image.new(
                'RGBA',
                (tpl_w, tpl_h),
                self.conf.get("general", "background_color")
            )
            
            img_cropped_w, img_cropped_h = self.img_cropped.size
            img_cropped_pos = (
                int(self.tpl_avatar_x - (img_cropped_w / 2)),
                int(self.tpl_avatar_y - (img_cropped_h / 2))
            )
            
            self.parse_user_info()
            
            if self.arg.verbose:
                self.img_cropped.save(
                    str(self.tmp_path / f"{self.user_id}.png"),
                    format="png"
                )
            
            self.bg_img.paste(self.img_cropped, img_cropped_pos)
            
            if self.arg.test:
                template_img.putalpha(125)
            
            # Use alpha_composite for proper alpha blending instead of paste with mask
            self.bg_img = Image.alpha_composite(self.bg_img, template_img)
            
            if self.arg.no_generate_qr:
                self.parse_qr_code()
            
            # Convert to RGB for text rendering (PIL text rendering works better with RGB)
            if self.bg_img.mode == 'RGBA':
                # Create white background and composite
                white_bg = Image.new('RGB', self.bg_img.size, 'white')
                white_bg.paste(self.bg_img, mask=self.bg_img.split()[3])  # Use alpha channel as mask
                self.bg_img = white_bg
            
            draw = ImageDraw.Draw(self.bg_img)
            # Calculate Y position with bounds check
            start_y = int(self.tpl_avatar_y + img_cropped_pos[1])
            # Ensure we don't start outside the image
            curr_y = max(0, min(start_y, self.tpl_h - 50))
            
            # Render username
            _, _, curr_y = self._render_text(draw, self.user_name, "username", curr_y)
            
            # Render position
            _, _, curr_y = self._render_text(draw, self.user_pos, "position", curr_y)
            
            # Render user ID
            userid_text = f"ID: {self.user_id}"
            _, _, _ = self._render_text(draw, userid_text, "userid", curr_y)
            
            # Save badge
            output_name = (
                f"{self.img_prefix}-"
                f"{self.user_name.upper()}_"
                f"{self.user_pos.upper()}_"
                f"{self.user_id.upper()}_"
                f"{self.img_num}.png"
            )
            self.bg_img.save(str(self.des_path / output_name), format="png")
            
        except IOError as error:
            LOGGER.error(f"IO Error: {error}")
            import traceback
            import sys
            traceback.print_exc(file=sys.stderr)
        except Exception as error:
            LOGGER.error(f"Unexpected error: {error}")
            import traceback
            import sys
            traceback.print_exc(file=sys.stderr)


def add_args(parser, action='exec'):
    """Add command-line arguments to parser"""
    if action in ['exec']:
        parser.add_argument(
            '-c', '--convert',
            action='store_true', default=False,
            help='Convert image to PNG format'
        )
        parser.add_argument(
            '--check-path',
            action='store_true', default=False,
            help='Check output path'
        )
        parser.add_argument(
            '-s', '--src-path',
            default=config.get("general", "src_path"),
            help='Path of the source folder'
        )
        parser.add_argument(
            '-f', '--des-path',
            default=config.get("general", "des_path"),
            help='Path of the destination folder'
        )
        parser.add_argument(
            '-t', '--template',
            default=config.get("template", "filename"),
            help='Template file name'
        )
        parser.add_argument(
            '-g', '--no-generate-qr',
            action='store_false', default=True,
            help='Skip generate QR code'
        )
        parser.add_argument(
            '-q', '--qr-text',
            help='QR code text'
        )
        parser.add_argument(
            '-a', '--no-auto-size',
            action='store_false', default=True,
            help='Auto size for text'
        )
        parser.add_argument(
            '-l', '--loop',
            type=bool, default=False,
            help='Loop the process'
        )
        parser.add_argument(
            '-i', '--interval',
            default=config.get('general', 'interval'),
            help='Interval for looping the process'
        )
        parser.add_argument(
            '--enable-crawler',
            action='store_true', default=False,
            help='Enable automatic image downloading if images are missing'
        )


def parse_cli() -> argparse.Namespace:
    """Parse command-line arguments"""
    parser = argparse.ArgumentParser(
        description=__doc__.strip(),
        formatter_class=Formatter
    )
    parser.add_argument('-d', '--debug', action='store_true')
    parser.add_argument('--test', action='store_true')
    parser.add_argument('-v', '--verbose', action='store_true')

    subparsers = parser.add_subparsers(help='Subcommand help', dest='action')
    add_args(
        subparsers.add_parser('exec', formatter_class=Formatter,
                              help='Full Execution'),
        'exec'
    )
    return parser.parse_args()


def setup_logging(debug: bool = False) -> None:
    """
    Setup logging configuration
    
    :param debug: Enable debug logging
    """
    log_level = logging.DEBUG if debug else logging.INFO
    logging.basicConfig(
        level=log_level,
        format='%(asctime)s %(levelname)-5s %(lineno)-4d %(funcName)s : %(message)s',
        stream=sys.stderr,
        filemode='w'
    )


def main(args: argparse.Namespace, config) -> None:
    """
    Main processing function
    
    :param args: Parsed command-line arguments
    :param config: Configuration object
    """
    files = []
    src_path = CUR_PATH / args.src_path
    des_path = CUR_PATH / args.des_path
    tmp_path = CUR_PATH / config.get("general", "tmp_path")
    cv_path = CUR_PATH / config.get("general", "converted_path")
    format_list = uT.get_list_file_extensions()

    if args.convert:
        ImageMaker.convert_images(str(src_path), str(cv_path))
        args.src_path = config.get("general", "converted_path")
        src_path = cv_path
    
    if args.check_path:
        Utility.check_folder([str(src_path), str(des_path), str(tmp_path)])

    # Collect all expected files from data source
    for r, d, f in os.walk(src_path):
        for file in f:
            if any(s for s in format_list if s in file):
                files.append(file)
    
    LOGGER.debug(f"Exec file list: {files}")
    
    # Initialize ImageManager for automatic image downloading if needed
    if hasattr(args, 'enable_crawler') and args.enable_crawler:
        LOGGER.info("Image crawler enabled - checking for missing images...")
        
        # Read employee data from Excel
        try:
            from openpyxl import load_workbook
            employee_ids = []
            excel_path = CUR_PATH / "data/employee_list.xlsx"
            if excel_path.exists():
                wb = load_workbook(excel_path)
                ws = wb.active
                for row in ws.iter_rows(values_only=True, min_row=2):
                    if len(row) >= 4:
                        uid, name, pos = row[1], row[2], row[3]
                        emp_id = f"{name}_{uid}_{pos}"
                        employee_ids.append(emp_id)
                LOGGER.info(f"Loaded {len(employee_ids)} employees from Excel")
            else:
                LOGGER.warning("Employee Excel file not found at data/employee_list.xlsx")
                employee_ids = files
        except Exception as e:
            LOGGER.warning(f"Error reading employee list: {e}, using existing files")
            employee_ids = files
        
        # Configure crawler URL from config or use default
        base_url = config.get("crawler", "base_url") if config.has_section("crawler") \
                   else "https://intranet.company.com/images/emp_images/big_new"
        
        img_manager = ImageManager(
            src_path=str(src_path),
            base_url=base_url,
            workers=int(config.get("crawler", "workers")) if config.has_section("crawler") else 5,
            timeout=30
        )
        
        # Download missing images using employee IDs
        download_stats = img_manager.download_missing_images(employee_ids, use_threading=True)
        
        LOGGER.info(f"Download Summary:")
        LOGGER.info(f"  Total files: {download_stats['total']}")
        LOGGER.info(f"  Missing: {download_stats['missing']}")
        LOGGER.info(f"  Downloaded: {download_stats['downloaded']}")
        LOGGER.info(f"  Failed: {download_stats['failed']}")
        LOGGER.info(f"  Already available: {download_stats['skipped']}")
        
        if download_stats['failed'] > 0:
            LOGGER.warning(f"Failed downloads: {download_stats['failed_files']}")
        
        # Re-collect files after download
        files.clear()
        for r, d, f in os.walk(src_path):
            for file in f:
                if any(s for s in format_list if s in file):
                    files.append(file)
        
        LOGGER.debug(f"Updated file list after download: {files}")

    count = 0
    while True:
        start = time.time()
        for file_name in files:
            count += 1
            LOGGER.info(f"Executing: {file_name}")
            img_maker = ImageMaker(file_name, args, config)
            img_maker.execute()
        end = time.time()
        LOGGER.info(f"Generated [{count} items] in [{end - start:.2f}] seconds...")
        if not args.loop:
            return
        Utility.countdown(int(args.interval))


if __name__ == "__main__":
    config = Utility.get_config()
    args = parse_cli()
    setup_logging(args.debug)
    main(args, config)
