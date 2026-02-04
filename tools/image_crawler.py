# -*- coding: utf-8 -*-
"""
Image crawler - download images from URLs using multi-threading.

Copyright (C) 2022
Authors: Vic Dang
Date: 22-Mar-22
Version: 1.0

Usage example:
  python image_crawler.py --workers 10 --file-path ./data.xlsx
"""

import argparse
import json
import logging
import logging.config
import queue
import socket
import sys
import threading
import urllib.request
import urllib.error
from pathlib import Path
from typing import List, Optional, Dict

from .util import Utilities as ut

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)


class ImageCrawler:
    """
    Multi-threaded image downloader from URLs.

    Downloads images from a web server based on employee IDs and stores them
    locally. Supports configurable worker threads for parallel downloads.
    """

    def __init__(
        self,
        workers: int = 10,
        base_url: str = "https://intranet.trna.com.vn/images/emp_images/big_new",
        file_type: int = 0,
        timeout: int = 30
    ) -> None:
        """
        Initialize ImageCrawler.

        Args:
            workers: Number of worker threads for parallel downloads.
            base_url: Base URL for downloading images.
            file_type: File type identifier (0: Excel format).
            timeout: Timeout in seconds for download requests.
        """
        self.workers = workers
        self.base_url = base_url
        self.file_type = file_type
        self.timeout = timeout
        self.failed_downloads: List[str] = []
        self.successful_downloads: int = 0

    def download_image(self, url: str, emp_id: str, output_path: Path) -> bool:
        """
        Download a single image from URL.

        Args:
            url: URL to download from.
            emp_id: Employee ID (format: name_id_position).
            output_path: Directory to save the image.

        Returns:
            True if download successful, False otherwise.
        """
        try:
            parts = emp_id.split('_')
            if len(parts) < 3:
                logger.warning(f"Invalid employee ID format: {emp_id}")
                return False

            name, uid, pos = parts[0], parts[1], parts[2]
            
            # Handle prefix (T/B)
            prep = ""
            if uid and uid[0] in ('T', 'B'):
                prep = uid[0]
                uid = uid[1:]
            
            # Convert UID to integer to remove leading zeros
            try:
                uid_int = int(uid)
            except ValueError:
                logger.error(f"Invalid UID format: {uid}")
                return False
            
            download_url = f"{url}/{uid_int}.webp"
            
            # Construct output filename
            output_file = output_path / f"{name}_{prep}{uid_int}_{pos}_1.webp"
            
            logger.debug(f"Downloading: {download_url}")
            
            # Use urlopen with timeout instead of urlretrieve (Python 3.14 compatible)
            with urllib.request.urlopen(download_url, timeout=self.timeout) as response:
                with open(str(output_file), 'wb') as out_file:
                    out_file.write(response.read())
            
            logger.info(f"Downloaded: {output_file}")
            self.successful_downloads += 1
            return True
            
        except urllib.error.HTTPError as err:
            logger.error(f"HTTP Error downloading {emp_id}: {err}")
            self.failed_downloads.append(emp_id)
            return False
        except urllib.error.URLError as err:
            logger.error(f"URL Error downloading {emp_id}: {err}")
            self.failed_downloads.append(emp_id)
            return False
        except Exception as err:
            logger.error(f"Failed to download {emp_id}: {err}")
            self.failed_downloads.append(emp_id)
            return False

    def _worker_thread(
        self,
        task_queue: queue.Queue,
        thread_no: int,
        output_path: Path
    ) -> None:
        """
        Worker thread function for processing download tasks.

        Args:
            task_queue: Queue containing download tasks.
            thread_no: Thread number identifier.
            output_path: Directory to save downloaded images.
        """
        while True:
            try:
                emp_id = task_queue.get(timeout=1)
                if emp_id is None:
                    break
                
                logger.debug(f"Thread {thread_no} processing: {emp_id}")
                self.download_image(self.base_url, emp_id, output_path)
                
            except queue.Empty:
                continue
            finally:
                task_queue.task_done()

    def start_workers(
        self,
        task_queue: queue.Queue,
        output_path: Path,
        num_workers: Optional[int] = None
    ) -> List[threading.Thread]:
        """
        Start worker threads for parallel downloading.

        Args:
            task_queue: Queue of download tasks.
            output_path: Directory to save downloaded images.
            num_workers: Number of worker threads (default: self.workers).

        Returns:
            List of started worker threads.
        """
        num_workers = num_workers or self.workers
        threads: List[threading.Thread] = []
        
        for i in range(num_workers):
            thread = threading.Thread(
                target=self._worker_thread,
                args=(task_queue, i, output_path),
                daemon=True
            )
            thread.start()
            threads.append(thread)
        
        return threads

    def download_batch(
        self,
        emp_ids: List[str],
        output_path: str
    ) -> Dict[str, int]:
        """
        Download multiple images using worker threads.

        Args:
            emp_ids: List of employee IDs to download.
            output_path: Directory to save images.

        Returns:
            Dictionary with statistics: {successful, failed, total}
        """
        output_dir = Path(output_path)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        task_queue: queue.Queue = queue.Queue()
        
        # Start worker threads
        threads = self.start_workers(task_queue, output_dir)
        
        # Queue all tasks
        for emp_id in emp_ids:
            emp_id = emp_id.strip()
            if emp_id:
                task_queue.put(emp_id)
        
        # Wait for queue to be processed
        task_queue.join()
        
        # Signal workers to stop
        for _ in range(len(threads)):
            task_queue.put(None)
        
        # Wait for all threads
        for thread in threads:
            thread.join()
        
        return {
            'successful': self.successful_downloads,
            'failed': len(self.failed_downloads),
            'total': len(emp_ids),
            'failed_ids': self.failed_downloads
        }


def setup_logging(debug: bool = False) -> None:
    """
    Setup logging configuration.

    Args:
        debug: Enable debug-level logging.
    """
    log_level = logging.DEBUG if debug else logging.INFO
    
    logging.basicConfig(
        level=log_level,
        format='%(asctime)s - %(levelname)s - %(threadName)s - '
               '%(name)s - %(funcName)s:%(lineno)d - %(message)s',
        stream=sys.stderr
    )


def parse_arguments() -> argparse.Namespace:
    """
    Parse command-line arguments.

    Returns:
        Parsed command-line arguments.
    """
    parser = argparse.ArgumentParser(
        description='Download images from web server using multi-threading.'
    )
    parser.add_argument(
        '-w', '--workers',
        type=int,
        default=10,
        help='Number of worker threads (default: 10)'
    )
    parser.add_argument(
        '-f', '--file-path',
        type=str,
        default="./data.xlsx",
        help='Path to file containing list of employee IDs'
    )
    parser.add_argument(
        '-u', '--url',
        type=str,
        default="https://intranet.trna.com.vn/images/emp_images/big_new",
        help='Base URL for downloading images'
    )
    parser.add_argument(
        '-o', '--output',
        type=str,
        default="./images/source/src_img",
        help='Output directory for downloaded images'
    )
    parser.add_argument(
        '-d', '--debug',
        action='store_true',
        help='Enable debug mode'
    )
    parser.add_argument(
        '-t', '--timeout',
        type=int,
        default=30,
        help='Timeout in seconds for downloads'
    )
    
    return parser.parse_args()


def get_data_from_file(file_path: str) -> List[str]:
    """
    Extract employee IDs from data file.

    Args:
        file_path: Path to data file (Excel or text).

    Returns:
        List of employee IDs.
        
    Raises:
        ImportError: If required module is not installed.
        FileNotFoundError: If data file not found.
    """
    data: List[str] = []
    file_type = ut.check_file_type(file_path)
    
    if file_type == "excel":
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(filename=file_path)
            sheet = workbook.active
            skip_first = True
            
            for row in sheet.iter_rows(values_only=True):
                if skip_first:
                    skip_first = False
                    continue
                
                if len(row) >= 4:
                    uid, name, pos = row[1], row[2], row[3]
                    emp_id = f"{name}_{uid}_{pos}"
                    data.append(emp_id)
        except ImportError as err:
            logger.error(f"Missing required module for Excel files: {err}")
            logger.error("Install openpyxl: pip install openpyxl")
            raise
        except FileNotFoundError:
            logger.error(f"Data file not found: {file_path}")
            raise
        except Exception as err:
            logger.error(f"Error reading Excel file: {err}")
            raise
    
    elif file_type == "txt":
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = [line.strip() for line in f if line.strip()]
        except FileNotFoundError:
            logger.error(f"Data file not found: {file_path}")
            raise
        except Exception as err:
            logger.error(f"Error reading text file: {err}")
            raise
    
    return data


def main() -> None:
    """Main execution function."""
    try:
        args = parse_arguments()
        setup_logging(args.debug)
        
        logger.info(f"Image Crawler Started")
        logger.info(f"Workers: {args.workers}, Timeout: {args.timeout}s")
        
        # Get employee IDs from file
        try:
            emp_ids = get_data_from_file(args.file_path)
        except Exception as err:
            logger.error(f"Fatal error reading data file: {err}")
            sys.exit(1)
        
        if not emp_ids:
            logger.warning(f"No employee IDs found in {args.file_path}")
            sys.exit(1)
        
        logger.info(f"Found {len(emp_ids)} employee IDs to process")
        
        # Download images
        try:
            crawler = ImageCrawler(
                workers=args.workers,
                base_url=args.url,
                timeout=args.timeout
            )
            
            stats = crawler.download_batch(emp_ids, args.output)
            
            logger.info(f"\n=== Download Summary ===")
            logger.info(f"Total: {stats['total']}")
            logger.info(f"Successful: {stats['successful']}")
            logger.info(f"Failed: {stats['failed']}")
            
            if stats['failed_ids']:
                logger.warning(f"Failed IDs: {stats['failed_ids']}")
                
            # Exit with 0 if at least some downloads succeeded
            if stats['successful'] > 0:
                sys.exit(0)
            else:
                logger.error("No images were successfully downloaded")
                sys.exit(1)
                
        except Exception as err:
            logger.error(f"Error during image download: {err}", exc_info=True)
            sys.exit(1)
            
    except Exception as err:
        logger.error(f"Fatal error in image crawler: {err}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
