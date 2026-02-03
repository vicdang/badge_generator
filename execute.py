# -*- coding: utf-8 -*-
# vim:ts=3:sw=3:expandtab
"""
 - Authors: Vic Dang
 - Skype: traxanh_dl
 - Usage example:
   + python execute.py -d -v exec
"""

import argparse
import glob
import logging.config
import os
import pathlib
import re
import sys
import time
import unicodedata
from argparse import ArgumentDefaultsHelpFormatter as Formatter
from pathlib import Path

# Add current directory to path to import tools package
sys.path.insert(0, str(Path(__file__).parent))

# Conditional imports - only load cv2/numpy if needed (not needed for cleanup)
try:
    import cv2
    import numpy as np
    HAS_CV2 = True
except ImportError:
    HAS_CV2 = False

try:
    import qrcode
    HAS_QRCODE = True
except ImportError:
    HAS_QRCODE = False

try:
    from PIL import Image, ImageDraw, ImageFont
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

from tools.util import Utilities as uT

logger = logging.getLogger()
PROJECT_ROOT = Path(__file__).parent.resolve()
CONFIG_FILE = PROJECT_ROOT / 'config' / 'config.ini'
config = None


class ImageSizeException(Exception):
   """ImageSizeException"""
   pass


class UserInfoException(Exception):
   """UserInfoException"""
   pass


class Utility(object):
   """docstring for Utility"""

   @staticmethod
   def get_config():
      try:
         from configparser import ConfigParser
      except ImportError:
         from ConfigParser import ConfigParser  # ver. < 3.0
      # instantiate
      conf = ConfigParser()
      # parse existing file
      conf.read(str(CONFIG_FILE))
      return conf

   @staticmethod
   def validate(string, regex):
      """
      Use for validating a string
      :param string: input string
      :param regex: regular expression to verify the string
      :return:
      """
      result = ""
      if re.search(regex, string):
         result = string
      else:
         logger.error("[{}] is not match [{}]".format(string, regex))
      return result

   @staticmethod
   def convert_img(input_img):
      """
      Convert image
      :param input_img: input image
      :return: converted image
      """
      if not HAS_CV2:
         logger.error("cv2 not available - cannot convert image")
         return None
      return cv2.imdecode(np.fromfile(input_img, dtype=np.uint8), -1)

   @staticmethod
   def countdown(due_time):
      """
      Used to countdown when processing
      :param due_time: due time
      """
      logger.debug("Counting down for %s" % due_time)
      while due_time:
         minute, second = divmod(due_time, 60)
         time_format = '{:02d}:{:02d}'.format(minute, second)
         print(time_format, end='\r')
         time.sleep(1)
         due_time -= 1

   @staticmethod
   def check_folder(folders):
      """
      Prepare required folders, including create folders and remove trunks
      :param folders: Target folder(s)
      """
      for folder in folders:
         if not os.path.isdir(folder):
            logger.debug("Making dir: %s" % folder)
            os.mkdir(folder)
         else:
            files = glob.glob(folder)
            for f in files:
               logger.debug("Removing: %s" % f)
               os.remove(f)

   @staticmethod
   def resize_with_aspect_ratio(image, width=None, height=None, inter=None):
      """
      resize_with_aspect_ratio
      :param image:
      :param width:
      :param height:
      :param inter:
      :return:
      """
      if not HAS_CV2:
         logger.error("cv2 not available - cannot resize image")
         return image
      
      if inter is None:
         inter = cv2.INTER_AREA
      
      (h, w) = image.shape[:2]
      if width is None and height is None:
         return image
      if width is None:
         r = height / float(h)
         dim = (int(w * r), height)
      else:
         r = width / float(w)
         dim = (width, int(h * r))

      return cv2.resize(image, dim, interpolation=inter)


class ImageMaker(object):
   """docstring for ImageMaker"""

   def __init__(self, name, arg, conf, image_full_path=None):
      super(ImageMaker, self).__init__()
      self.arg = arg
      self.name = name
      self.image_full_path = image_full_path  # Store full path if provided
      self.src_path = str(PROJECT_ROOT / arg.src_path)
      self.des_path = str(PROJECT_ROOT / arg.des_path)
      self.tmp_path = str(PROJECT_ROOT / conf.get("general", "tmp_path"))
      self.template = str(PROJECT_ROOT / arg.template)
      self.debug = arg.debug or False
      self.img_prefix = conf.get("general", "img_prefix")
      self.tpl_avatar_x = conf.getint("template", "avata_x")
      self.tpl_avatar_y = conf.getint("template", "avata_y")
      self.tpl_avatar_w = conf.getint("template", "avata_w")
      self.tpl_avatar_h = conf.getint("template", "avata_h")
      self.tpl_w = conf.getint("template", "width")
      self.tpl_h = conf.getint("template", "height")
      self.conf = conf
      self.base_text_size = self.conf.getint("general", "text_size")
      self.positions = uT.get_dict_positions()
      self.re_name = r"^[\w\.\- ]+$"
      self.re_id = r"^[\w]?[\d]+$"

      self.img = self.img_resized = self.img_cropped = self.bg_img = self.qr_img = None
      self.user_pos = self.positions["E"]
      self.user_name = self.user_id = ""
      self.img_num = 1

   def get_focus_position(self, input_img):
      """
      Used to detect faces in the image then return the focus position
      :param input_img: The input image
      :return: Focus position of the image (x, y)
      """
      if not HAS_CV2:
         logger.warning("cv2 not available, using default focus position (center)")
         return (0, 0)  # Return center as fallback
      
      face_cascade = cv2.CascadeClassifier(
            str(PROJECT_ROOT / 'resources' / 'haar_cascade' / 'haarcascade_frontalface_default.xml'))
      image = Utility.convert_img(input_img)
      gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
      faces = face_cascade.detectMultiScale(gray_image,
                                            # scaleFactor=1.2,
                                            scaleFactor=self.conf.getfloat(
                                                  "avata", "scalefactor"),
                                            minNeighbors=3,
                                            minSize=(30, 30),
                                            maxSize=(200, 200),
                                            flags=cv2.CASCADE_SCALE_IMAGE)

      logger.info("Found {0} faces!".format(len(faces)))
      x = y = 0
      # when face(s) detected
      if len(faces):
         # Track the base position (of the very first face)
         a, b, c, d = faces[0]
         for (i, j, w, h) in faces:
            logger.debug("Detected face : [{},{},{},{}]".format(i, j, w, h))
            c = int((c + w) / 2)
            d = int((d + h) / 2)
            a = int((a + i) / 2)
            b = int((b + j) / 2)
            x = a + int(c / 2)
            y = b + int(d / 2)
            if self.arg.debug:
               cv2.rectangle(image,
                             (i + int(w / 2), j + int(h / 2)),
                             (i + int(w / 2) + 2, j + int(h / 2) + 2),
                             (0, 200, 100), 2)
               cv2.rectangle(image, (i, j), (i + w, j + h), (0, 200, 100), 2)
               cv2.rectangle(image, (x, y), (x + 2, y + 2), (0, 255, 255), 2)
      if self.debug:
         resize = Utility.resize_with_aspect_ratio(image, height=1024)
         cv2.imshow("Faces found", resize)
         cv2.imwrite(os.path.join(self.tmp_path, "faces_" + str(self.name)),
                     image)
         if self.arg.test and self.arg.verbose:
            cv2.waitKey(0)
      logger.debug("Focus coordinate : [{},{}]".format(x, y))
      return x, y

   @staticmethod
   def convert_images(src_path, des_path):
      """
      Convert Images
      :param src_path: Source path
      :param des_path: Destination path
      """
      # Ensure source directory exists
      if not os.path.exists(src_path):
         logger.error(f"Source path does not exist: {src_path}")
         return
      
      # Create destination directory if it doesn't exist
      os.makedirs(des_path, exist_ok=True)
      
      files = []
      format_list = uT.get_list_file_extensions()
      
      try:
         f = [fi for fi in os.listdir(src_path) if os.path.isfile(os.path.join(src_path, fi))]
         logger.debug("File List : %s" % [item for item in f])
         for file in f:
            if any(s for s in format_list if s in file):
               files.append(u'{}'.format(file))
         
         logger.info(f"Converting {len(files)} images to PNG format...")
         for f in files:
            name = f.split(".")[0]
            file_name = os.path.join(src_path, f)
            try:
               logger.info("Converting : %s to PNG format ..." % file_name)
               im = Image.open(file_name).convert("RGBA")
               im.save(os.path.join(des_path, name + '.png'), format="png")
            except Exception as e:
               logger.error(f"Failed to convert {file_name}: {e}")
      except Exception as err:
         logger.error(f"Error during image conversion: {err}")

   @staticmethod
   def cleanup_images(clean_root, skip_paths):
      """
      Recursively cleanup all files in clean_root directory, excluding skip_paths
      :param clean_root: Root directory to clean (relative to PROJECT_ROOT)
      :param skip_paths: List of directories to skip (protect from deletion)
      :return: Count of deleted files
      """
      deleted_count = 0
      clean_full_path = str(PROJECT_ROOT / clean_root)
      
      if not os.path.exists(clean_full_path):
         logger.error(f"Cleanup root path does not exist: {clean_full_path}")
         return 0
      
      logger.info(f"Cleaning directory recursively: {clean_full_path}")
      logger.info(f"Protected paths: {skip_paths}")
      
      # Normalize skip paths to full paths for comparison (ensure exact matching, not substring)
      skip_full_paths = [str(PROJECT_ROOT / skip_path).replace('\\', '/').rstrip('/') 
                         for skip_path in skip_paths]
      
      # Walk through the directory tree (bottom-up for safe deletion)
      for root, dirs, files in os.walk(clean_full_path, topdown=False):
         # Convert current path to forward slashes for consistent comparison
         current_path_normalized = root.replace('\\', '/')
         
         # Check if current directory should be skipped
         # Use exact path matching: either the dir is exactly a skip_path, or it's under a skip_path
         should_skip = False
         for skip_path in skip_full_paths:
            skip_normalized = skip_path.rstrip('/')
            curr_normalized = current_path_normalized
            # Exact match or child of skip path (using path separator to avoid substring false positives)
            if curr_normalized == skip_normalized or curr_normalized.startswith(skip_normalized + '/'):
               should_skip = True
               logger.debug(f"Skipping protected directory: {root}")
               break
         
         if should_skip:
            continue
         
         # Delete files in this directory
         for file in files:
            file_path = os.path.join(root, file)
            try:
               os.remove(file_path)
               logger.debug(f"Deleted: {file_path}")
               deleted_count += 1
            except Exception as e:
               logger.error(f"Failed to delete {file_path}: {e}")
         
         # Try to remove empty directories (but not skip_paths and not essential directories like source, output, temp)
         essential_dirs = ['source', 'output', 'temp', 'templates', 'cv']
         for dir_name in dirs:
            dir_path = os.path.join(root, dir_name)
            dir_path_normalized = dir_path.replace('\\', '/')
            
            # Never delete essential directories
            if any(essential in dir_path_normalized for essential in essential_dirs):
               logger.debug(f"Skipping essential directory: {dir_path}")
               continue
            
            # Check if directory should be skipped (exact path matching)
            should_skip_dir = False
            for skip_path in skip_full_paths:
               skip_normalized = skip_path.rstrip('/')
               # Exact match or child of skip path (using path separator to avoid substring false positives)
               if dir_path_normalized == skip_normalized or dir_path_normalized.startswith(skip_normalized + '/'):
                  should_skip_dir = True
                  logger.debug(f"Skipping protected directory: {dir_path}")
                  break
            
            if should_skip_dir:
               continue
            
            try:
               if os.path.isdir(dir_path) and not os.listdir(dir_path):  # Directory is empty
                  os.rmdir(dir_path)
                  logger.debug(f"Removed empty directory: {dir_path}")
            except Exception as e:
               logger.debug(f"Could not remove directory {dir_path}: {e}")
      
      return deleted_count

   def correct_img(self):
      """
      Correct the image
      """
      orientation = 274  # get 274 through upper loop
      try:
         # Handle both old and new Pillow versions
         if hasattr(self.img, '_getexif'):
            exif = self.img._getexif()
         else:
            try:
               from PIL.Image import Exif
               exif = self.img.getexif()
            except (ImportError, AttributeError):
               exif = None
         if exif:
            exif = dict(exif.items())
            if exif[orientation] == 3:
               self.img.rotate(180, expand=True)
            elif exif[orientation] == 6:
               self.img.rotate(270, expand=True)
            elif exif[orientation] == 8:
               self.img.rotate(90, expand=True)
      except Exception as err:
         logger.error(err)
         # There is AttributeError: _getexif sometimes.
         pass

   def resize_img(self):
      """
      Resize image
      :return:
      """
      img_w, img_h = self.img.size
      logger.info("Width: %f, Height: %f" % (img_w, img_h))
      basewidth = self.tpl_avatar_w + self.conf.getint("template", "padding")
      if img_h >= img_w:
         wpercent = (basewidth / float(img_w))
         hsize = int((float(img_h) * float(wpercent)))
         self.img_resized = self.img.resize((basewidth, hsize),
                                            Image.LANCZOS)
      else:
         hpercent = (basewidth / float(img_h))
         wsize = int((float(img_w) * float(hpercent)))
         self.img_resized = self.img.resize((wsize, basewidth),
                                            Image.LANCZOS)
      # cropped_example = cropped_example.resize((basewidth,hsize),
      # Image.LANCZOS)
      logger.debug("Resized image: %f x %f" % self.img_resized.size)
      self.img_resized.save(os.path.join(self.tmp_path, self.name), format="png")

   def crop_img(self):
      """
      Crop image
      """
      basewidth = self.tpl_avatar_w + self.conf.getint("template", "padding")
      x, y = self.get_focus_position(os.path.join(self.tmp_path, self.name))
      img_r_w, img_r_h = self.img_resized.size
      if img_r_w < basewidth or img_r_h < basewidth:
         raise ImageSizeException("Image is too small, Please try another.")
      base_w = basewidth / 2
      if x <= 0 and y <= 0:
         x = int(img_r_w / 2)
         y = int(img_r_h / 2)
      else:
         if x - base_w <= 0:
            x = base_w
         elif x + base_w > img_r_w:
            x = img_r_w - base_w
         if y - base_w <= 0:
            y = base_w
         elif y + base_w > img_r_h:
            y = base_w
      correct_x, correct_y = x - base_w, y - base_w
      correct_w, correct_h = basewidth + correct_x, basewidth + correct_y
      logger.debug("[x:{}, y:{}] - [c_x:{}, c_y:{}] - [w:{}, h:{}]".format(
            x, y, correct_x, correct_y, basewidth, basewidth))
      draw = ImageDraw.Draw(self.img_resized)
      if self.debug:
         draw.rectangle([correct_x, correct_y, correct_w, correct_h], width=3,
                        outline="#0000ff")
      self.img_cropped = self.img_resized.crop((correct_x, correct_y,
                                                correct_w, correct_h))
      logger.debug("Cropped image: %f x %f" % self.img_cropped.size)
      if self.arg.debug or self.arg.test:
         self.img_cropped.save(os.path.join(self.tmp_path, "cr_" + self.name), format="png")

   def parse_user_info(self):
      """
      Parse user information
      Parse user information
      :return: 
      """
      img_info_arr = self.name.split(".")[0].split("_")
      self.user_name = Utility.validate(img_info_arr[0].strip(),
                                        self.re_name)
      if self.user_name == 0:
         raise UserInfoException("User name not found!")
      self.user_pos = img_info_arr[2].capitalize()
      if self.user_pos.strip().upper() in self.positions:
         self.user_pos = self.positions[self.user_pos.strip().upper()]
         logger.info("pos: {}".format(self.user_pos))
      elif self.user_pos.strip().upper() in list(map(str.upper, 
                                                     self.positions.values())):
         self.user_pos = self.user_pos.strip().upper()
         logger.info("pos: {}".format(self.user_pos))
      else:
         logger.error(
               "[{}] is not in [{}]".format(self.user_pos, self.positions))
         raise UserInfoException("User position is incorrect!")
      self.user_id = Utility.validate(img_info_arr[1].strip(), self.re_id)
      if self.user_id == 0:
         raise UserInfoException("User ID not found!")
      if len(img_info_arr) == 4:
         self.img_num = img_info_arr[3]

   def parse_qr_code(self):
      """
      Make QR code
      """
      self.qr_img = None  # Initialize
      
      if not HAS_QRCODE:
         logger.warning("qrcode not available, skipping QR code generation")
         return  # Skip QR code if qrcode module not available
      
      try:
         # Make RQ code
         qr = qrcode.QRCode(version=self.conf.getint("qrcode", "version"),
                            error_correction=qrcode.constants.ERROR_CORRECT_L,
                            box_size=self.conf.getint("qrcode", "boxsize"),
                            border=self.conf.getint("qrcode", "border"))
         name = unicodedata.normalize('NFKD', self.user_name).encode('ascii',
                                                                     'ignore')
         img_info = "Fullname: %s,Position: %s, Badge_Id: %s, Company: %s" % (
            name, self.user_pos, self.user_id, "https://www.tma.vn")
         logger.info("info: {}".format(img_info))
         if self.arg.qr_text:
            self.qr_img = qrcode.make(self.arg.qr_text)
         else:
            qr.add_data(img_info)
            qr.make(fit=self.conf.get("qrcode", "fit"))
            self.qr_img = qr.make_image(fill_color=self.conf.get("qrcode",
                                                            "fillcolor"),
                                 back_color=self.conf.get("qrcode",
                                                            "backcolor"))
         
         # Ensure QR image is in RGBA format for consistent pasting
         if self.qr_img and self.qr_img.mode != 'RGBA':
            self.qr_img = self.qr_img.convert('RGBA')
         
         qr_x = self.conf.getint("qrcode", "qr_x")
         qr_y = self.conf.getint("qrcode", "qr_y")
         qr_w = self.conf.getint("qrcode", "qr_w")
         qr_h = self.conf.getint("qrcode", "qr_h")
         
         if self.qr_img:
            # Resize QR code to fit the specified dimensions
            qr_resized = self.qr_img.resize((qr_w, qr_h), Image.Resampling.LANCZOS)
            # Use 4-tuple bounding box (left, top, right, bottom) for paste
            self.bg_img.paste(qr_resized, (qr_x, qr_y, qr_x + qr_w, qr_y + qr_h), qr_resized)
      except Exception as e:
         logger.error("QR code generation failed: {}".format(str(e)))
         self.qr_img = None  # Ensure qr_img is None on error

   def execute(self):
      """
      Execute image maker
      :return:
      """
      try:
         # Ensure destination directory exists
         os.makedirs(self.des_path, exist_ok=True)
         
         logger.info("Processing Image: %s" % self.name)
         template_img = Image.open(self.template, 'r')
         tpl_w, tpl_h = template_img.size

         # Use the full path if provided (for images in subdirectories), else construct from src_path + name
         if self.image_full_path:
            image_path = self.image_full_path
         else:
            image_path = os.path.join(self.src_path, self.name)
         
         self.img = Image.open(image_path, 'r').convert("RGBA")
         curr_y = 0.0
         self.correct_img()
         self.resize_img()
         self.crop_img()
         # Make background
         self.bg_img = Image.new('RGBA', (tpl_w, tpl_h), self.conf.get(
               "general", "background_color"))
         img_cropped_w, img_cropped_h = self.img_cropped.size
         img_cropped_pos = (int(self.tpl_avatar_x - (img_cropped_w / 2)),
                            int(self.tpl_avatar_y - (img_cropped_h / 2)))
         self.parse_user_info()
         if self.arg.verbose:
            self.img_cropped.save(os.path.join(self.tmp_path, self.user_id + ".png"),
                                  format="png")
         self.bg_img.paste(self.img_cropped, img_cropped_pos)
         if self.arg.test:
            template_img.putalpha(125)
         self.bg_img.paste(template_img, (0, 0), mask=template_img)
         if self.arg.no_generate_qr:
            self.parse_qr_code()
         draw = ImageDraw.Draw(self.bg_img)

         curr_y += self.tpl_avatar_y + img_cropped_pos[1]
         th = self.conf.getint("username", "toppad")
         if self.user_name:
            curr_y += th
            if self.arg.no_auto_size:
               size = self.conf.getint("username", "size")
            else:
               size = self.base_text_size
            n_size = size
            while True:
               font = ImageFont.truetype(os.path.join(str(PROJECT_ROOT),
                                                      self.conf.get(
                                                            "username",
                                                            "font")), n_size)
               msg = self.user_name.upper().strip()
               bbox = draw.textbbox((0, 0), msg, font=font)
               tw = bbox[2] - bbox[0]
               th = bbox[3] - bbox[1]
               logger.info("info: {} {} {}".format(msg, tw, th))
               if tw < (tpl_w - 50):
                  break
               else:
                  n_size = n_size - 5
            draw.text(((tpl_w - tw) / 2, curr_y), msg, self.conf.get(
                  "username", "color"), font=font)

         if self.user_pos:
            curr_y += th + self.conf.getint("position", "toppad")
            if self.arg.no_auto_size:
               size = self.conf.getint("position", "size")
            else:
               size = self.base_text_size - 10
            font = ImageFont.truetype(os.path.join(str(PROJECT_ROOT), self.conf.get(
                  "position",
                  "font")), size)
            msg = self.user_pos.strip()
            bbox = draw.textbbox((0, 0), msg, font=font)
            tw = bbox[2] - bbox[0]
            th = bbox[3] - bbox[1]
            draw.text(((tpl_w - tw) / 2, curr_y), msg, self.conf.get(
                  "position", "color"), font=font)

         if self.user_id:
            curr_y += th + self.conf.getint("userid", "toppad")
            if self.arg.no_auto_size:
               size = self.conf.getint("userid", "size")
            else:
               size = self.base_text_size - 15
            font = ImageFont.truetype(os.path.join(str(PROJECT_ROOT), self.conf.get(
                  "userid",
                  "font")), size)
            msg = "ID: " + self.user_id.strip()
            bbox = draw.textbbox((0, 0), msg, font=font)
            tw = bbox[2] - bbox[0]
            draw.text(((tpl_w - tw) / 2, curr_y), msg,
                      self.conf.get("userid", "color"), font=font)

         # Saved in the same relative location
         self.bg_img.save(os.path.join(self.des_path, self.img_prefix + "-" +
                          self.user_name.upper() + "_" +
                          self.user_pos.upper() + "_" +
                          self.user_id.upper() + "_" +
                          str(self.img_num) + ".png"), format="png")

      except IOError as error:
         logger.error("Error: %s" % error)


def add_args(parser, action='exec'):
   """
   :param parser:
   :param action:
   """
   if action in ['exec']:
      parser.add_argument('-c', '--convert',
                          action='store_true', default=False,
                          help='Convert image to PNG format')
      parser.add_argument('--check-path',
                          action='store_true', default=False,
                          help='Check out put')
      parser.add_argument('-s', '--src-path',
                          default=config.get("general", "src_path"),
                          help='Path of the source folder')
      parser.add_argument('-f', '--des-path',
                          default=config.get("general", "des_path"),
                          help='Path of the destination folder')
      parser.add_argument('-t', '--template',
                          default=config.get("template", "filename"),
                          help='Template file name')
      parser.add_argument('-g', '--no-generate-qr',
                          action='store_false', default=True,
                          help='Skip generate QR code')
      parser.add_argument('-q', '--qr-text',
                          help='RQ code text')
      parser.add_argument('-a', '--no-auto-size',
                          action='store_false', default=True,
                          help='Auto size for text')
      parser.add_argument('-l', '--loop',
                          type=bool, default=False,
                          help='Lopping the process')
      parser.add_argument('-i', '--interval',
                          default=config.get('general', 'interval'),
                          help='Interval for looping the process')


def parse_cli():
   """
   :return:
   """
   parser = argparse.ArgumentParser(description=__doc__.strip(),
                                    formatter_class=argparse
                                    .ArgumentDefaultsHelpFormatter)
   parser.add_argument('-d', '--debug', action='store_true')
   parser.add_argument('--test', action='store_true')
   parser.add_argument('-v', '--verbose', action='store_true')

   subparsers = parser.add_subparsers(help='Subcommand help', dest='action')
   add_args(subparsers.add_parser('exec', formatter_class=Formatter,
                                  help='Full Execution'), 'exec')
   subparsers.add_parser('cleanup', formatter_class=Formatter,
                         help='Cleanup image directories')
   return parser.parse_args()


def setup_logging(debug=False):
   """
   :param debug:
   """
   if debug:
      log_level = logging.DEBUG
   else:
      log_level = logging.INFO
   logging.basicConfig(level=log_level,
                       format='%(asctime)s %(levelname)-5s %(lineno)-4d '
                              '%(funcName)s : %(message)s',
                       stream=sys.stderr,
                       filemode='w')
   global logger
   logger = logging.getLogger('sLogger')


def main(args, config):
   """
   Main processing
   :return:
   """
   # Handle cleanup action
   if hasattr(args, 'action') and args.action == 'cleanup':
      if not config.has_section('cleanup') or not config.getboolean('cleanup', 'enabled'):
         logger.error("Cleanup is not configured or disabled in config.ini")
         return
      
      clean_root = config.get('cleanup', 'clean_root')
      skip_paths_str = config.get('cleanup', 'skip_paths')
      skip_paths = [p.strip() for p in skip_paths_str.split(',') if p.strip()]
      
      logger.info(f"Starting cleanup of: {clean_root}")
      logger.info(f"Protected paths: {skip_paths}")
      
      deleted_count = ImageMaker.cleanup_images(clean_root, skip_paths)
      logger.info(f"Cleanup completed! Deleted {deleted_count} files.")
      return
   
   # Original exec logic - only proceed if action is exec
   if not hasattr(args, 'src_path'):
      logger.error("No action specified. Use 'exec' or 'cleanup'")
      return
   
   files = []
   src_path = str(PROJECT_ROOT / args.src_path)
   des_path = str(PROJECT_ROOT / args.des_path)
   tmp_path = str(PROJECT_ROOT / config.get("general", "tmp_path"))
   cv_path = str(PROJECT_ROOT / config.get("general", "converted_path"))
   format_list = uT.get_list_file_extensions()

   if args.check_path:
      Utility.check_folder([src_path, des_path, tmp_path])

   # Try to load employee data from Excel file first
   data_file = PROJECT_ROOT / 'data' / 'employee_list.xlsx'
   employees_data = {}
   
   if data_file.exists():
      logger.info(f"Loading employee data from {data_file}")
      try:
         from openpyxl import load_workbook
         workbook = load_workbook(filename=str(data_file))
         sheet = workbook.active
         skip_first = True
         
         # Parse Excel: Column 0=ID, Column 1=UID, Column 2=Name, Column 3=Position
         for row in sheet.iter_rows(values_only=True):
            if skip_first:
               skip_first = False
               continue
            if len(row) >= 4:
               row_id, emp_uid, name, position = row[0], row[1], row[2], row[3]
               if emp_uid and name:
                  # Create employee ID in format: Name_UID_Position for crawler
                  emp_id = f"{str(name).strip()}_{str(emp_uid).strip()}_{str(position).strip() if position else ''}"
                  employees_data[emp_id] = {
                     'name': str(name).strip(),
                     'position': str(position).strip() if position else "",
                     'uid': str(emp_uid).strip()
                  }
         logger.info(f"Loaded {len(employees_data)} employees from Excel")
      except Exception as err:
         logger.error(f"Error loading Excel file: {err}")
   
   # Discover available images in source folder
   available_images = {}
   files = []
   for root, dirs, file_list in os.walk(src_path):
      for file in file_list:
         if any(s for s in format_list if s in file):
            full_path = os.path.join(root, file)
            files.append(file)
            available_images[file] = full_path
   
   logger.debug("Available images: %s" % [item for item in files])
   
   # If we have employee data, try to match images with data
   if employees_data:
      logger.info(f"Matching {len(files)} images with {len(employees_data)} employees")
      # Try to crawl missing images
      try:
         from tools.image_crawler import ImageCrawler
         crawler_config = {
            'base_url': config.get('crawler', 'base_url'),
            'workers': config.getint('crawler', 'workers'),
            'timeout': config.getint('crawler', 'timeout')
         }
         logger.info(f"Attempting to crawl missing images from {crawler_config['base_url']}")
         
         missing_ids = [emp_id for emp_id in employees_data.keys() 
                       if not any(emp_id in img for img in files)]
         if missing_ids:
            logger.info(f"Found {len(missing_ids)} employees with missing images, crawling...")
            crawler = ImageCrawler(**crawler_config)
            crawler.download_batch(missing_ids, src_path)
            
            # Re-discover available images after crawling
            available_images = {}
            files = []
            for root, dirs, file_list in os.walk(src_path):
               for file in file_list:
                  if any(s for s in format_list if s in file):
                     full_path = os.path.join(root, file)
                     files.append(file)
                     available_images[file] = full_path
            logger.info(f"After crawling: {len(files)} images available")
         else:
            logger.info("All employees have images")
      except Exception as err:
         logger.warning(f"Could not crawl missing images: {err}")
   
   logger.debug("Final file list: %s" % [item for item in files])
   
   # NOW convert images to PNG if requested (after crawling)
   if args.convert:
      logger.info("Converting images to PNG format...")
      ImageMaker.convert_images(src_path, cv_path)
      args.src_path = config.get("general", "converted_path")
      src_path = cv_path
      
      # Re-discover available images after conversion, but only PNG files
      available_images = {}
      files = []
      for root, dirs, file_list in os.walk(src_path):
         for file in file_list:
            if file.lower().endswith('.png'):
               full_path = os.path.join(root, file)
               files.append(file)
               available_images[file] = full_path
      logger.info(f"After conversion: {len(files)} PNG images available")
   
   count = 0
   while True:
      start = time.time()
      for file_name in files:
         count += 1
         # Get full path from available_images dict
         full_path = available_images.get(file_name)
         if not full_path:
            # Fallback: search for the file in subdirectories
            for root, dirs, file_list in os.walk(src_path):
               if file_name in file_list:
                  full_path = os.path.join(root, file_name)
                  break
         if not full_path:
            # Last resort: try direct path (for backward compatibility)
            full_path = os.path.join(src_path, file_name)
         
         logger.info("Executing: %s" % file_name)
         img_maker = ImageMaker(file_name, args, config, image_full_path=full_path)
         img_maker.execute()
      end = time.time()
      logger.info("Generated [" + str(count) + " items] in [" + str(end -
                                                                    start) +
                  "] seconds...")
      if not args.loop:
         return
      Utility.countdown(int(args.interval))


if __name__ == "__main__":
   config = Utility.get_config()
   args = parse_cli()
   setup_logging(args.debug)
   main(args, config)
   # try:
   #    logging.debug('Execute with arguments :')
   #    logging.debug(args)
   #    main(args, config)
   # except Exception as err:
   #    logging.error('Error performing action: %s, %s' % (args.action, err))
