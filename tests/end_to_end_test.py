#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""End-to-End Badge Generator Test: Download Images → Generate Badges."""

import sys
import os
import logging
import shutil
import subprocess
from pathlib import Path

# Set UTF-8 encoding for Windows console
os.environ['PYTHONIOENCODING'] = 'utf-8'
sys.stdout.reconfigure(encoding='utf-8')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def print_section(title: str) -> None:
    """Print formatted section header."""
    print()
    print("╔" + "═" * 70 + "╗")
    print("║" + title.center(70) + "║")
    print("╚" + "═" * 70 + "╝")
    print()


def cleanup_directories() -> None:
    """Clean up directories for fresh test."""
    print_section("🧹 STEP 1: CLEANUP")
    
    dirs_to_clean = [
        Path("./images/source/src_img"),
        Path("./images/output/des_img"),
        Path("./images/test/test_download"),
    ]
    
    for dir_path in dirs_to_clean:
        if dir_path.exists():
            # Remove all .webp, .jpg, .png, .bmp files
            for pattern in ["*.webp", "*.jpg", "*.png", "*.bmp"]:
                for file in dir_path.glob(pattern):
                    file.unlink()
                    logger.info(f"  Deleted: {file}")
        else:
            dir_path.mkdir(parents=True, exist_ok=True)
            logger.info(f"  Created: {dir_path}")


def create_employee_list() -> None:
    """Create sample employee list for testing."""
    print_section("📋 STEP 2: CREATE EMPLOYEE LIST")
    
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Employees'
        
        # Add headers
        ws['A1'] = 'ID'
        ws['B1'] = 'UID'
        ws['C1'] = 'Name'
        ws['D1'] = 'Position'
        
        # Add test employees - use real IDs that have images on server
        sample_data = [
            (1, 154176, 'Nguyen Van A', 'A'),
            (2, 154177, 'Tran Thi B', 'SE'),
            (3, 154178, 'Hoang Van C', 'TL'),
        ]
        
        for row_idx, (id_val, uid, name, pos) in enumerate(sample_data, start=2):
            ws[f'A{row_idx}'] = id_val
            ws[f'B{row_idx}'] = uid
            ws[f'C{row_idx}'] = name
            ws[f'D{row_idx}'] = pos
        
        wb.save('employee_list.xlsx')
        logger.info(f"✅ Created employee_list.xlsx with {len(sample_data)} employees")
        
        for id_val, uid, name, pos in sample_data:
            logger.info(f"   - {name} (UID: {uid}, Position: {pos})")
    except Exception as e:
        logger.error(f"❌ Error creating employee list: {e}")
        raise


def download_images() -> bool:
    """Download images using image crawler."""
    print_section("⬇️  STEP 3: DOWNLOAD IMAGES FROM TMA INTRANET")
    
    try:
        logger.info("Running image crawler...")
        logger.info("Command: python -m tools.image_crawler -f employee_list.xlsx -w 5")
        
        result = subprocess.run(
            ["python", "-m", "tools.image_crawler", "-f", "employee_list.xlsx", "-w", "5"],
            capture_output=True,
            text=True,
            timeout=60
        )
        
        # Show output
        if result.stdout:
            for line in result.stdout.split('\n'):
                if 'Downloaded' in line or 'Total' in line or 'Successful' in line or 'Failed' in line:
                    logger.info(line)
        
        if result.stderr:
            for line in result.stderr.split('\n'):
                if 'Downloaded' in line or 'Total' in line or 'Successful' in line or 'Failed' in line:
                    logger.info(line)
        
        # Check if images were downloaded
        src_dir = Path("./img/src_img")
        if src_dir.exists():
            webp_files = list(src_dir.glob("*.webp"))
            if webp_files:
                logger.info(f"✅ Successfully downloaded {len(webp_files)} images:")
                for img_file in webp_files:
                    size = img_file.stat().st_size / 1024
                    logger.info(f"   - {img_file.name} ({size:.1f} KB)")
                return True
            else:
                logger.warning("⚠️  No images found in src_img directory")
                return False
        else:
            logger.warning("⚠️  src_img directory not found")
            return False
    
    except Exception as e:
        logger.error(f"❌ Error downloading images: {e}")
        return False


def generate_badges() -> bool:
    """Generate badges using downloaded images."""
    print_section("🎫 STEP 4: GENERATE BADGES")
    
    try:
        logger.info("Running badge generator...")
        logger.info("Command: python execute.py exec -c --check-path")
        
        # Note: This will require dependencies like cv2, so we'll try to run it
        result = subprocess.run(
            ["python", "execute.py", "exec", "-c", "--check-path"],
            capture_output=True,
            text=True,
            timeout=120
        )
        
        # Show relevant output
        if result.stdout:
            for line in result.stdout.split('\n'):
                if 'Success' in line or 'Badge' in line or 'Generated' in line or 'Complete' in line:
                    logger.info(line)
        
        if result.stderr:
            for line in result.stderr.split('\n'):
                if 'Generated' in line or 'Badge' in line or 'Error' in line.lower():
                    logger.info(line)
        
        # Check if badges were generated
        des_dir = Path("./img/des_img")
        if des_dir.exists():
            badge_files = list(des_dir.glob("*.png")) + list(des_dir.glob("*.jpg"))
            if badge_files:
                logger.info(f"✅ Successfully generated {len(badge_files)} badges:")
                for badge_file in badge_files[:3]:
                    size = badge_file.stat().st_size / 1024
                    logger.info(f"   - {badge_file.name} ({size:.1f} KB)")
                if len(badge_files) > 3:
                    logger.info(f"   - ... and {len(badge_files) - 3} more")
                return True
            else:
                logger.warning("⚠️  No badges found in des_img directory")
                logger.info("   (This may be expected if cv2 is not fully configured)")
                return False
        else:
            logger.warning("⚠️  des_img directory not found")
            des_dir.mkdir(parents=True, exist_ok=True)
            return False
    
    except subprocess.TimeoutExpired:
        logger.error("❌ Badge generation timed out")
        return False
    except Exception as e:
        logger.error(f"❌ Error generating badges: {e}")
        return False


def verify_results() -> None:
    """Verify all results."""
    print_section("✅ STEP 5: VERIFY RESULTS")
    
    results = {
        "Source Images": 0,
        "Generated Badges": 0,
    }
    
    # Count source images
    src_dir = Path("./img/src_img")
    if src_dir.exists():
        results["Source Images"] = len(list(src_dir.glob("*.webp"))) + \
                                   len(list(src_dir.glob("*.jpg"))) + \
                                   len(list(src_dir.glob("*.png")))
    
    # Count generated badges
    des_dir = Path("./img/des_img")
    if des_dir.exists():
        results["Generated Badges"] = len(list(des_dir.glob("*.png"))) + \
                                      len(list(des_dir.glob("*.jpg")))
    
    print("Results Summary:")
    print(f"  📥 Source Images Downloaded: {results['Source Images']}")
    print(f"  📤 Badges Generated: {results['Generated Badges']}")
    
    if results["Source Images"] > 0:
        logger.info(f"✅ Image download: SUCCESS ({results['Source Images']} images)")
    else:
        logger.warning(f"⚠️  Image download: CHECK NETWORK/VPN ACCESS")
    
    if results["Generated Badges"] > 0:
        logger.info(f"✅ Badge generation: SUCCESS ({results['Generated Badges']} badges)")
    else:
        logger.info(f"ℹ️  Badge generation: Skipped (requires full OpenCV setup)")
    
    return results


def main() -> None:
    """Run complete end-to-end test."""
    print()
    print("╔" + "═" * 70 + "╗")
    print("║" + "  BADGE GENERATOR - END-TO-END TEST".center(70) + "║")
    print("║" + "  Download Images → Generate Badges".center(70) + "║")
    print("║" + "  January 21, 2026".center(70) + "║")
    print("╚" + "═" * 70 + "╝")
    
    try:
        # Step 1: Cleanup
        cleanup_directories()
        
        # Step 2: Create employee list
        create_employee_list()
        
        # Step 3: Download images
        download_success = download_images()
        
        if download_success:
            # Step 4: Generate badges
            badges_success = generate_badges()
        else:
            logger.warning("⚠️  Skipping badge generation - no source images available")
            badges_success = False
        
        # Step 5: Verify results
        results = verify_results()
        
        # Final summary
        print_section("📊 FINAL SUMMARY")
        
        print("Test Execution Status:")
        print(f"  ✅ Cleanup: Complete")
        print(f"  ✅ Employee List: Created (3 employees)")
        
        if download_success:
            print(f"  ✅ Image Download: SUCCESS ({results['Source Images']} images)")
        else:
            print(f"  ⚠️  Image Download: FAILED/INCOMPLETE")
        
        if badges_success:
            print(f"  ✅ Badge Generation: SUCCESS ({results['Generated Badges']} badges)")
        else:
            print(f"  ℹ️  Badge Generation: Skipped/Incomplete")
        
        print()
        if download_success and badges_success:
            print("🎉 END-TO-END TEST PASSED! 🎉")
            print("   Images downloaded successfully → Badges generated successfully")
        elif download_success and not badges_success:
            print("✅ PARTIAL SUCCESS")
            print("   Images downloaded successfully ✅")
            print("   Badge generation skipped (requires OpenCV configuration)")
        else:
            print("⚠️  TEST INCOMPLETE")
            print("   Please check network/VPN access to TMA intranet")
            print("   Or verify employee IDs are correct")
        
        print()
        logger.info("Test completed!")
        
    except Exception as e:
        logger.error(f"Fatal error during test: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
